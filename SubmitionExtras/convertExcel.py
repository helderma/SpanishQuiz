from openpyxl import workbook as wb
from openpyxl import load_workbook

from sys import stdout
def formatData(chunks):    
    fullData =""
        
    for x in xrange(len(chunks)):
        #if a number is found then convert it to string
        if type(chunks[x]) == type(long()) or type(chunks[x]) == type(float()) or type(chunks[x]) == type(int()):
            fullData += "\""+ str(chunks[x])+"\""
        #otherwise format the data to be useful for CSV format
        
        else:
            fullData += "\""+chunks[x]+"\""
            '''
            chunks[x].decode('utf-8')
            for xx in xrange(len(chunks[x])):
                if chunks[x][xx] == ",":
                    chunks[x][xx] = chunks[x][xx].decode('utf-8')
                    chunks[x][xx] = "/"
            '''
        if(x != len(chunks)-1):
            fullData += ","
    #fullData +=","
    
    return fullData

def retriveBookData(path):
    workbook = load_workbook(path)
    sheets = workbook.sheetnames
    #access all sheets in a workbook
    sheetIndex =0
    bookData = list()
    for s in sheets:
        sheet = workbook[s]
        #acess all rows, includes empties
        #print sheet.title
        bookData.append(sheet.title)
        sheetData = list()
        for row in sheet.iter_rows():
            #access all cells in a row
            cells = row
            #create a list to save rows
            rowData = list()        
            #access a single cell in a row
            for cell in cells:           
                #access cells who have a value
                if cell.value != None:
                    rowData.append(cell.value)
            # only access data if it's not empty
            if rowData:
                sheetData.append(formatData(rowData))
        bookData.append(sheetData)
            
    return bookData
def saveToCSVs(info):
    folderPath ="./csv/"
    f = ""
    for x in xrange(len(info)):
        if x%2 == 0:
            f = open(folderPath+info[x]+".csv",'w')
            print info[x]
        else:
            for l in info[x]:
                #print l
                #line = bytearray(l.decode('utf-8'))
                for ch in l:
                    f.write(str(ch.encode('utf-8')))
                f.write('\n')


path = 'testing.xlsx'
data = retriveBookData(path)
saveToCSVs(data)
raw_input('waiting to end')
#ws = workbook.get_sheet_by_name(sheets[0])

#print ws.rows
#file = open(,'w')

    
'''
def readRows(row):
    for 
'''